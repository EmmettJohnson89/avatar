from selenium import webdriver
from selenium.common.exceptions import *
import json
import func
import os
from time import sleep
import pandas as pd
import openpyxl

def submit():
    with open('verified_emails.json', 'r') as file:
        accounts = json.load(file)
    for account in accounts:
        if account['applied'] == False:
            name = account['profile']
            email = account['email']
            password = account['password']

            with open(f'profiles/{name}.json', 'r') as fp:
                profile = json.load(fp)

            driver = webdriver.Chrome()
            driver.maximize_window()
            action = webdriver.ActionChains(driver)

            func.login(driver, email, password)
            func.get_started(driver)
            
            account['applied'] = True
            with open('verified_emails.json', 'w') as file:
                json.dump(accounts, file)

            func.select_experience(driver)
            func.select_what_is_my_goal(driver)
            func.select_work_preference(driver)
            func.select_manualmode(driver)
            func.add_professional(driver, profile['professional'])
            func.add_experience(driver, profile['work_experience'])
            func.add_education(driver, profile['education'], action)
            func.add_language(driver, profile['languages'])
            func.add_skills(driver, profile['skills'])
            func.add_overview(driver, profile['overview'])
            func.add_service(driver, profile['services'])
            func.add_rate(driver, profile['hour_rate'])
            func.add_photo_others(driver, profile['photo_others'], action)
            func.submit_profile(driver)
            func.notification_imme(driver)
            func.agency_name(driver, profile['agency'])
            driver.quit()

if __name__ == "__main__":
    response_1 = input('Would you like to create new emails? (y/n) :')
    response_2 = input('Would you like to apply your profile to new emails? (y/n) :')
    # response_1 = 'y'
    # response_2 = 'y'
    if response_1 == 'y':
        count = int(input('Enter number of accounts for each profile. '))
        if count > 0:
            subfix = '.'
            while True:
                text = input(f'Enter profiles to be used{subfix} ')
                profiles = text.split(' ')
                exists = True
                for name in profiles:
                    if not os.path.exists(f'profiles/{name}.json'):
                        exists = False
                if exists:
                    break
                else:
                    subfix = ' again.'
                sleep(1)
            profiles = text.split(' ')
            excel = openpyxl.load_workbook("email.xlsx")
            sheet = excel.active
            excel_file = "email.xlsx"
            rowCount = sheet.max_row + 1
            emails = []
            for row in range(1, rowCount):
                email = sheet.cell(row=row, column=1).value
                emails.append({"name" : email, "verified" : False})
            print(emails)
            # emails = func.get_email(count)
            func.verify_email(emails, profiles, count)
    if response_2 == 'y':
        submit()
    print('done')